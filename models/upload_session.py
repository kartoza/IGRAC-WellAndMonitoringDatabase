# coding=utf-8
"""Upload session model definition.

"""
import gc
import json
import ntpath
import os
import uuid
from datetime import datetime, timedelta

import openpyxl
from celery import current_app
from celery.result import AsyncResult
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import models
from django.db.models import Q
from django.utils.timezone import now, make_aware
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font

from gwml2.models.metadata.license_metadata import LicenseMetadata
from gwml2.models.well import Well
from gwml2.models.well_management.organisation import Organisation
from gwml2.utilities import ods_to_xlsx, xlsx_to_ods

UPLOAD_SESSION_CATEGORY_WELL_UPLOAD = 'well_upload'
UPLOAD_SESSION_CATEGORY_MONITORING_UPLOAD = 'well_monitoring_upload'
UPLOAD_SESSION_CATEGORY_DRILLING_CONSTRUCTION_UPLOAD = (
    'well_drilling_and_construction'
)

# make choices
UPLOAD_SESSION_CATEGORY = (
    (
        UPLOAD_SESSION_CATEGORY_WELL_UPLOAD,
        UPLOAD_SESSION_CATEGORY_WELL_UPLOAD
    ),
    (
        UPLOAD_SESSION_CATEGORY_MONITORING_UPLOAD,
        UPLOAD_SESSION_CATEGORY_MONITORING_UPLOAD
    ),
    (
        UPLOAD_SESSION_CATEGORY_DRILLING_CONSTRUCTION_UPLOAD,
        UPLOAD_SESSION_CATEGORY_DRILLING_CONSTRUCTION_UPLOAD
    )
)

User = get_user_model()


class TaskStatus:
    """Task status."""

    RUNNING = 'RUNNING'
    STOP = 'STOP'
    DONE = 'DONE'


class UploadSessionCancelled(Exception):
    def __init__(self, error):
        super(Exception, self).__init__(error)
        self.errors = error


class UploadSessionCheckpoint:
    """Upload session checkpoint."""

    SAVING_DATA = 'Saving data'
    CACHE_WELLS = 'Cache wells'
    CACHE_COUNTRY = 'Cache country'
    CACHE_ORGANISATION = 'Cache organisation'
    CREATE_REPORT = 'Create report'
    FINISH = 'Finish'

    STEP_CHECKPOINT_CHOICES = [
        (1, SAVING_DATA),
        (2, CACHE_WELLS),
        (3, CACHE_COUNTRY),
        (4, CACHE_ORGANISATION),
        (5, CREATE_REPORT),
        (6, FINISH),
    ]

    @staticmethod
    def get_index(checkpoint):
        """Get index."""
        LABEL_TO_VALUE = {
            label: value for value, label in
            UploadSessionCheckpoint.STEP_CHECKPOINT_CHOICES
        }
        return LABEL_TO_VALUE.get(checkpoint)


class UploadSession(LicenseMetadata):
    """Upload session model
    """
    organisation = models.ForeignKey(
        Organisation,
        on_delete=models.SET_NULL,
        null=True,
        blank=True
    )

    token = models.UUIDField(
        default=uuid.uuid4,
        editable=False,
        null=True
    )

    uploader = models.IntegerField(
        null=True,
        blank=True
    )

    uploaded_at = models.DateTimeField(
        default=datetime.now
    )

    category = models.CharField(
        choices=UPLOAD_SESSION_CATEGORY,
        max_length=50,
        blank=True,
        default='',
    )

    step = models.TextField(
        blank=True,
        null=True
    )

    status = models.TextField(
        help_text=(
            'What is the status of progress. '
            'Mostly it will split by added, error and skipped ',
        ),
        blank=True,
        null=True
    )

    progress = models.IntegerField(
        default=0
    )

    upload_file = models.FileField(
        upload_to='gwml2/upload/',
        null=True
    )

    is_processed = models.BooleanField(
        default=False
    )

    is_canceled = models.BooleanField(
        default=False
    )

    task_id = models.TextField(
        blank=True,
        null=True
    )

    # Upload method
    is_adding = models.BooleanField(
        default=True, help_text='Does the upload add new data.'
    )
    is_updating = models.BooleanField(
        default=False, help_text='Does the update existing data.'
    )

    # Checkpoints
    checkpoint = models.IntegerField(
        choices=UploadSessionCheckpoint.STEP_CHECKPOINT_CHOICES,
        default=UploadSessionCheckpoint.get_index(
            UploadSessionCheckpoint.SAVING_DATA
        )
    )
    checkpoint_ids = models.JSONField(
        blank=True,
        null=True,
        help_text=(
            'This is the ids for the checkpoint, '
            'e.g: list of wells id for CACHE_COUNTRY checkpoint'
        )
    )

    # noinspection PyClassicStyleClass
    class Meta:
        """Meta class for project."""
        verbose_name_plural = 'Upload sessions'
        verbose_name = 'Upload session'
        ordering = ('-uploaded_at',)
        db_table = 'upload_session'

    def __str__(self):
        return str(self.token)

    @property
    def filename(self):
        return ntpath.basename(self.upload_file.name)

    @property
    def timestamp(self):
        return self.uploaded_at.timestamp() * 1000

    def get_uploader(self):
        """ return user of uploader """
        try:
            return User.objects.get(id=self.uploader)
        except User.DoesNotExist:
            return None

    def progress_status(self):
        """
        return progress status
        if there is file of progress
        else return from model value itself
        """
        try:
            _file = open(os.path.join(
                settings.MEDIA_ROOT, 'gwml2', 'upload', 'progress',
                str(self.token)
            ))
            return json.loads(_file.read())
        except Exception:
            return {
                'progress': self.progress,
                'status': self.status
            }

    def check_cancellation(self):
        """Check cancellation."""
        self.refresh_from_db()
        if self.is_canceled:
            raise UploadSessionCancelled('Cancelled')

    def update_progress(self, finished=False, status='', progress=0):
        """Update progress for current upload session

        :param finished: Whether the session is finished or not
        :type finished: bool

        :param status: Status message for current progress
        :type status: str

        :param progress: Current progress
        :type progress: int
        """
        self.check_cancellation()

        if finished:
            if progress == 100:
                self.is_processed = True
            if progress < 100:
                self.is_canceled = True

        self.progress = progress
        if status:
            self.status = status
        self.save()

    def update_status(self, sheet_name, status):
        """Update status."""
        self.check_cancellation()
        try:
            _status = json.loads(self.status)
        except Exception:
            _status = {}

        _status[sheet_name] = status
        self.status = json.dumps(_status)
        self.save()

    def update_step(self, step: str, progress: int = None):
        """Update step of upload."""
        self.check_cancellation()
        self.step = step
        if progress:
            self.progress = progress
        self.save()

    @property
    def task_status(self):
        """Return task status from celery."""
        # If processed, meaning done
        if self.is_processed:
            return TaskStatus.DONE
        if self.task_id:
            if self.is_canceled:
                return TaskStatus.STOP

            active_tasks = current_app.control.inspect().active()
            if not active_tasks:
                return TaskStatus.STOP
            for worker, running_tasks in active_tasks.items():
                for task in running_tasks:
                    if task["id"] == self.task_id:
                        return TaskStatus.RUNNING

            active_tasks = current_app.control.inspect().reserved()
            for worker, running_tasks in active_tasks.items():
                for task in running_tasks:
                    if task["id"] == self.task_id:
                        return TaskStatus.RUNNING

        time_difference = now() - (
            self.uploaded_at if self.uploaded_at.tzinfo else make_aware(
                self.uploaded_at
            )
        )

        if time_difference > timedelta(minutes=10):
            self.is_canceled = True
            self.save()

        return TaskStatus.STOP

    def stop(self):
        """Stop the progress."""
        if self.task_id:
            AsyncResult(self.task_id, app=current_app).revoke(terminate=True)
        self.is_canceled = True
        self.save()

    def run_in_background(self, restart: bool = False):
        """Run the uploader in background."""
        from gwml2.tasks import well_batch_upload
        if self.task_status != TaskStatus.RUNNING:
            well_batch_upload.delay(self.id, restart)

    def run(self, restart: bool = False):
        """Run the upload."""
        from gwml2.tasks.uploader.uploader import BatchUploader
        from gwml2.tasks.uploader import (
            DrillingAndConstructionUploader,
            GeneralInformationUploader,
            HydrogeologyUploader,
            ManagementUploader,
            MonitoringDataUploader,
            StratigraphicLogUploader,
            StructuresUploader,
            WaterStrikeUploader
        )
        if self.category == UPLOAD_SESSION_CATEGORY_WELL_UPLOAD:
            BatchUploader(
                self,
                [
                    GeneralInformationUploader, HydrogeologyUploader,
                    ManagementUploader
                ],
                restart
            )
        elif self.category == UPLOAD_SESSION_CATEGORY_MONITORING_UPLOAD:
            BatchUploader(self, [MonitoringDataUploader], restart)
        elif self.category == (
                UPLOAD_SESSION_CATEGORY_DRILLING_CONSTRUCTION_UPLOAD
        ):
            BatchUploader(
                self,
                [
                    DrillingAndConstructionUploader, WaterStrikeUploader,
                    StratigraphicLogUploader, StructuresUploader
                ],
                restart
            )

    @property
    def file_report_url(self):
        """Return URL for file report upload."""
        _url = self.upload_file.url
        ext = os.path.splitext(_url)[1]
        return _url.replace(ext, f'.report{ext}')

    def create_report_excel(self):
        """Created excel that will contain reports."""
        self.step = 'Create report'
        self.save()
        try:
            _file = ods_to_xlsx(self.upload_file.path)

            # Create file report name
            ext = os.path.splitext(_file)[1]
            _report_file = _file.replace(ext, f'.report{ext}')

            # If file report equals file, skip create report
            if _report_file == _file:
                return
            if os.path.exists(_report_file):
                os.remove(_report_file)
            workbook = openpyxl.load_workbook(_file)

            query = self.uploadsessionrowstatus_set.order_by(
                'row'
            ).exclude(status=0)
            status_column = {}
            for sheetname in workbook.sheetnames:
                sheet_query = query.filter(
                    Q(sheet_name=sheetname) | Q(
                        sheet_name=sheetname.replace('_', ' '))
                )
                worksheet = workbook[sheetname]

                # We need to check latest column
                try:
                    status_column_idx = status_column[sheetname]
                except KeyError:
                    row = worksheet[1]
                    status_column_idx = len(row) + 1
                    last_column = False
                    for _row in row:
                        if isinstance(_row, MergedCell):
                            continue
                        if _row.value:
                            status_column_idx = _row.column + 1
                        elif not last_column:
                            status_column_idx = _row.column
                            last_column = True

                    status_column[sheetname] = status_column_idx

                for idx, row_status in enumerate(sheet_query):
                    row_status.update_sheet(worksheet, status_column_idx)
            workbook.save(_report_file)
            workbook.close()
            os.chmod(_report_file, 0o0777)
            xlsx_to_ods(_report_file)

            # Delete files
            del workbook
            gc.collect()
            os.remove(_file)
            os.remove(_report_file)
            self.step = 'Create report done'
            self.save()
        except Exception as e:
            print(f'{e}')

    @staticmethod
    def running_sessions():
        """Return running session status."""
        return UploadSession.objects.filter(
            is_processed=False,
            is_canceled=False,
            progress__lt=100
        )

    def resume(self):
        """Resume the upload."""
        if self.task_status == TaskStatus.STOP:
            self.run_in_background()

    def append_checkout_ids(self, value):
        """Append checkout ids."""
        if not self.checkpoint_ids:
            self.checkpoint_ids = []
        self.checkpoint_ids.append(value)
        self.save()


RowStatus = [
    (0, 'Added'),
    (1, 'Error'),
    (2, 'Skipped')
]

ADDED_FILL = PatternFill(
    start_color='00FF00', end_color='00FF00', fill_type='solid'
)
ERROR_FILL = PatternFill(
    start_color='FF0000', end_color='FF0000', fill_type='solid'
)
SKIPPED_FILL = PatternFill(
    start_color='FFFF00', end_color='FFFF00', fill_type='solid'
)
ERROR_FONT = Font(color="FFFFFF")


class UploadSessionRowStatus(models.Model):
    """ Check status data per row of upload """
    upload_session = models.ForeignKey(UploadSession, on_delete=models.CASCADE)
    sheet_name = models.CharField(max_length=256)
    row = models.IntegerField()
    column = models.IntegerField()
    status = models.IntegerField(choices=RowStatus)
    note = models.TextField(
        null=True, blank=True
    )
    well = models.ForeignKey(
        Well, on_delete=models.SET_NULL, null=True, blank=True
    )

    class Meta:
        verbose_name_plural = 'Upload sessions row statuses'
        verbose_name = 'Upload session row status'
        db_table = 'upload_session_row_status'
        unique_together = ['upload_session', 'sheet_name', 'row', 'column']

    def update_sheet(self, worksheet, status_column_idx):
        """Update the sheet."""
        try:
            cell = worksheet.cell(
                row=self.row, column=self.column + 1
            )
            status = ''
            if self.status == 0:
                status = 'Added'
                cell.fill = ADDED_FILL
            elif self.status == 1:
                status = 'Error'
                cell.fill = ERROR_FILL
                cell.font = ERROR_FONT
                worksheet.cell(
                    row=self.row, column=self.column + 1
                ).value = f'Error: {self.note}'
            elif self.status == 2:
                status = 'Skipped'
                cell.fill = SKIPPED_FILL
            worksheet.cell(
                row=self.row, column=status_column_idx
            ).value = status
        except Exception as e:
            print(f'{e}')
            pass
