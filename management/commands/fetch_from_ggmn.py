# NOTE : To run this command you need to install reverse_geocoder
# pip install reverse_geocoder
# coding=utf-8
import ast
import requests
from requests.auth import HTTPBasicAuth
import re
import json
import os
import time
import calendar
import csv
from datetime import datetime
from shutil import copyfile
from openpyxl import load_workbook
import reverse_geocoder as rg

from django.conf import settings
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model

from gwml2.tasks.uploader.well import (
    create_or_get_well, create_monitoring_data
)
from gwml2.models.well_management.organisation import Organisation

User = get_user_model()

CODE = 'code'
NAME = 'name'
GEOMETRY = 'geometry'
COORDINATES = 'coordinates'
LAT = 'lat'
LON = 'lon'
STATION_TYPE = 'station_type'
WELL_TEMPLATE_FILE = 'wells.xlsx'
WELL_MONITORING_FILE = 'monitoring_data.xlsx'
STATUS = 'status'
SURFACE_LEVEL = 'surface_level'
TIME_SERIES = 'timeseries'
FILTERS = 'filters'

# -- MONITORING
PARAMETER_GWmBGS = 'Water depth [from the ground surface]'
PARAMETER_GWmMSL = 'Water level elevation a.m.s.l.'
PARAMETER = {
    'GWmBGS': PARAMETER_GWmBGS,
    'GWmMSL': PARAMETER_GWmMSL
}

COUNTRIES_CODE_FILE = 'countries_codes_and_coordinates.csv'
ORGANISATION_NAME = 'ggmn'

DJANGO_ROOT = os.path.dirname(
    os.path.dirname(
        os.path.dirname(os.path.abspath(__file__))
    ))


class Command(BaseCommand):
    ggmn_station_url = 'https://demo.lizard.net/api/v3/groundwaterstations/'
    ggmn_timeseries_url = 'https://demo.lizard.net/api/v3/timeseries/'
    ggmn_organisation_url = 'https://demo.lizard.net/api/v3/organisations/'
    well_ids = []
    well_data = {}
    countries_code = {}
    max_page = 0
    admin_id = None
    organisation = None
    username = ''
    password = ''
    command_list = [
        'organisations', 'wells', 'location'
    ]

    def add_arguments(self, parser):
        parser.add_argument(
            '-u',
            '--username',
            dest='username',
            default='',
            help='GGMN account username')
        parser.add_argument(
            '-p',
            '--password',
            dest='password',
            default='',
            help='GGMN account password')
        parser.add_argument(
            '-m',
            '--max_page',
            dest='max_page',
            default=0,
            help='Max pages to fetch, default is fetch all')
        parser.add_argument(
            '-f',
            '--from_page',
            dest='from_page',
            default=1,
            help='Start the command from this page index, default = 1')
        parser.add_argument(
            '-db',
            '--upload-to-db',
            dest='upload_to_db',
            default='False'
        )
        parser.add_argument(
            '-c',
            '--commands',
            dest='commands',
            default='organisation,wells,location'
        )

    def get_current_timestamp(self):
        """
        Return current timestamp
        """
        gmt = time.gmtime()
        return calendar.timegm(gmt)

    def read_countries_code(self):
        """
        Read data from countries code csv,
        then store it as a dictionary
        """
        csv_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            COUNTRIES_CODE_FILE
        )
        with open(csv_path) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                line_count += 1
                if line_count == 1:
                    continue
                self.countries_code[
                    row[1].strip().replace('"', '')
                ] = row[2].strip().replace('"', '')

    def fetch_organisations(self, page=1):
        """Fetch organisations data then store them to database"""
        organisation_url = '{base_url}?page={page}'.format(
            base_url=self.ggmn_organisation_url,
            page=page
        )
        print(organisation_url)
        response = requests.get(
            organisation_url,
            auth=(self.username, self.password))
        organisations = response.json()
        for organisation in organisations['results']:
            organisation_obj, _ = Organisation.objects.get_or_create(
                name=organisation['name'],
                description='organisation from ggmn'
            )
            print(organisation_obj.name, _)
            if 'users_url' in organisation:
                print('## Fetching users')
                print(organisation['users_url'])
                users_response = requests.get(
                    organisation['users_url'],
                    auth=(self.username, self.password)
                )
                users = users_response.json()
                for user in users:
                    user_obj, user_created = (
                        User.objects.get_or_create(
                            username=user['username'],
                            first_name=user['first_name'],
                            last_name=user['last_name'],
                            email=user['email']
                        )
                    )
                    if user_obj.id not in organisation_obj.viewers:
                        organisation_obj.viewers.append(user_obj.id)
                        organisation_obj.save()
                        print('User {username} has been added as a viewer to {org}'.format(
                            username=user_obj.username,
                            org=organisation_obj.name
                        ))

        if 'next' in organisations and organisations['next']:
            next_url = organisations['next']
            next_page = int(re.findall(r'\d+', next_url)[-1])
            self.fetch_organisations(next_page)


    def fetch_wells(self, well_template_output_path, page=1):
        """
        Fetch wells from api
        """
        if self.max_page:
            if page > self.max_page:
                return

        print('#- Processing page {}'.format(page))
        well_url = '{base_url}?page={page}'.format(
            base_url=self.ggmn_station_url,
            page=page
        )
        print(well_url)
        response = requests.get(well_url)
        ground_stations = response.json()

        if well_template_output_path:
            wb = load_workbook(well_template_output_path)
            ws = wb.worksheets[0]
        else:
            wb, ws = None, None

        for ground_station in ground_stations['results']:
            if ground_station[CODE] in self.well_ids:
                continue
            self.well_ids.append(ground_station[CODE])
            print('#-- Processing well {}'.format(ground_station[CODE]))
            lat = ground_station[GEOMETRY][COORDINATES][1]
            lon = ground_station[GEOMETRY][COORDINATES][0]
            try:
                location = rg.search(
                    (lat, lon),
                )
            except TypeError:
                continue

            # Store well data for fetching monitoring data
            self.well_data[ground_station[CODE]] = []
            if FILTERS in ground_station:
                filters = ground_station[FILTERS]
                if len(filters) > 0:
                    _filter = filters[0]
                    if TIME_SERIES in _filter:
                        for time_series_data in _filter[TIME_SERIES]:
                            self.well_data[ground_station[CODE]].append({
                                time_series_data['name']:
                                    time_series_data['uuid']
                            })
            try:
                row_data = [
                    ground_station[CODE],  # ID
                    ground_station[NAME],  # NAME
                    ground_station[STATUS].capitalize(),  # Status
                    '',  # Feature Type
                    '',  # Purpose
                    lat,  # Lat
                    lon,  # Lon
                    ground_station[SURFACE_LEVEL],  # Ground surface elevation value
                    'm',  # Ground surface elevation unit
                    '', '',  # Elevation of the top of the casing ( unknown )
                    self.countries_code[
                        location[0]['cc'].upper()],
                    '',  # Address
                    ''  # Description
                ]
                if well_template_output_path:
                    ws.append(row_data)
                else:
                    # Upload to db directly
                    if not self.organisation:
                        self.organisation, _ = Organisation.objects.get_or_create(
                            name=ORGANISATION_NAME
                        )
                    if not self.admin_id:
                        self.admin_id = User.objects.filter(
                            is_superuser=True,
                            username='admin'
                        )[0].id
                    well, created = create_or_get_well(
                        organisation=self.organisation,
                        data=row_data,
                        additional_data={
                            'created_by': self.admin_id,
                            'last_edited_by': self.admin_id,
                            'description': json.dumps(
                                self.well_data[ground_station[CODE]])
                        }
                    )
                    print('Well {name}, Status : {status}'.format(
                        name=well.original_id,
                        status='Created' if created else 'Exist'
                    ))
            except KeyError:
                pass
        if wb:
            wb.save(well_template_output_path)
        if 'next' in ground_stations and ground_stations['next']:
            next_url = ground_stations['next']
            next_page = int(re.findall(r'\d+', next_url)[-1])
            if self.max_page:
                if next_page <= self.max_page:
                    self.fetch_wells(well_template_output_path, next_page)
            else:
                self.fetch_wells(well_template_output_path, next_page)

    def fetch_monitoring_data(self, template_output_path, ):
        """
        Fetch monitoring data for wells
        """
        today_timestamp = int(time.mktime(datetime.today().timetuple()) * 1000)

        if not self.well_data:
            return
        if template_output_path:
            wb = load_workbook(template_output_path)
            ws = wb.worksheets[0]
        else:
            wb, ws = None, None

        # self.well_data e.g. {'B10H0082':
        # [{"GWmMSL": "167761f9-7691-44fa-a11e-addf331c3ff8"},
        # {"GWmBGS": "258f8f51-6c97-4829-ade3-b1ac4642ef72"}],..}
        for well_code in self.well_data:
            # well_code => "B10H0082"
            # well_data => [
            # {"GWmMSL": "167761f9-7691-44fa-a11e-addf331c3ff8"},
            # {"GWmBGS": "258f8f51-6c97-4829-ade3-b1ac4642ef72"}]
            well_data = self.well_data[well_code]

            # time_series => {"GWmMSL": "167761f9-7691-44fa-a11e-addf331c3ff8"}
            for time_series in well_data:
                for time_series_name in time_series:
                    # uuid => "167761f9-7691-44fa-a11e-addf331c3ff8"
                    # time_series_name => "GWmMSL"
                    uuid = time_series[time_series_name]
                    print('#-- Processing {well} - {name} uuid {uuid}'.format(
                        well=well_code,
                        name=time_series_name,
                        uuid=uuid))
                    monitoring_url = (
                        '{base_url}?min_points={min_points}&start={start}'
                        '&end={end}'
                        '&uuid={uuid}'
                    ).format(
                        base_url=self.ggmn_timeseries_url,
                        min_points=0,
                        start=0,
                        end=today_timestamp,
                        uuid=uuid
                    )
                    print(monitoring_url)
                    response = requests.get(monitoring_url)
                    monitoring_data = response.json()

                    results = monitoring_data['results']
                    for result in results:
                        events = result['events']
                        for event in events:
                            timestamp = int(event['timestamp'] / 1000)
                            date_object = datetime.fromtimestamp(timestamp)
                            row_data = [
                                well_code,  # ID
                                date_object.strftime('%Y-%m-%d %H:%M:%S'),  # Date and time
                                PARAMETER[time_series_name],  # Parameter
                                event['value'],  # Value
                                'm',  # Unit
                                '',  # Method
                            ]
                            if template_output_path:
                                ws.append(row_data)
                            else:
                                create_monitoring_data(
                                    organisation_name=self.organisation.name,
                                    data=row_data,
                                    additional_data={
                                        'created_by': self.admin_id,
                                        'last_edited_by': self.admin_id
                                    },
                                )

        if template_output_path:
            wb.save(template_output_path)

    def handle(self, *args, **options):
        """Implementation for command.
        :param args:  Not used
        :param options: Not used

        """
        self.read_countries_code()
        self.max_page = int(options.get('max_page'))
        from_page = int(options.get('from_page'))
        if self.max_page > 0:
            self.max_page +=  from_page - 1
        upload_to_db = ast.literal_eval(
            options.get('upload_to_db', 'False')
        )

        commands = options.get('commands', '').split(',')
        self.username = options.get('username', '')
        self.password = options.get('password', '')

        if 'organisations' in commands:
            print('# Fetching organisations')
            self.fetch_organisations()

        if 'wells' in commands:
            print('# Fetching Wells ')
            if not upload_to_db:
                well_template_path = os.path.join(
                    settings.STATIC_ROOT,
                    'download_template',
                    WELL_TEMPLATE_FILE
                )
                well_template_output_name = '{file}_{time}.xlsx'.format(
                    file='WELL',
                    time=self.get_current_timestamp()
                )
                well_template_output_path = os.path.join(
                    DJANGO_ROOT,
                    well_template_output_name
                )
                copyfile(
                    well_template_path,
                    well_template_output_path
                )
            else:
                well_template_output_path = None
            self.fetch_wells(well_template_output_path, page=from_page)

            print('# Fetching Monitoring data')
            if not upload_to_db:
                well_monitoring_template_path = os.path.join(
                    settings.STATIC_ROOT,
                    'download_template',
                    WELL_MONITORING_FILE
                )
                well_monitoring_template_output_name = '{file}_{time}.xlsx'.format(
                    file='WELL_MONITORING',
                    time=self.get_current_timestamp()
                )
                well_monitoring_template_output_path = os.path.join(
                    DJANGO_ROOT,
                    well_monitoring_template_output_name
                )
                copyfile(
                    well_monitoring_template_path,
                    well_monitoring_template_output_path
                )
            else:
                well_monitoring_template_output_path = None
            self.fetch_monitoring_data(well_monitoring_template_output_path)
