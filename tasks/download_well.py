import json
import os
import shutil
import time
import zipfile

from django.contrib.auth import get_user_model
from django.conf import settings
from django.http import JsonResponse
from shutil import copyfile
from openpyxl import load_workbook

from celery import shared_task
from celery.utils.log import get_task_logger
from gwml2.models.download_session import DownloadSession
from gwml2.models.well import Well

logger = get_task_logger(__name__)

User = get_user_model()


def filter_wells_to_download(filters):
    wells = Well.objects.all()
    if not filters:
        return wells

    # feature_type filter
    feature_type_data = filters.get('feature_type', None)
    if feature_type_data:
        value = feature_type_data.get('value', '')
        operator = feature_type_data.get('operator', '')
        if value and operator:
            if operator == 'ilike':
                wells = wells.filter(feature_type__name__icontains=value)
            elif operator == '=':
                wells = wells.filter(feature_type__name__iexact=value)

    # original_id filter
    original_id_data = filters.get('original_id', None)
    if original_id_data:
        value = original_id_data.get('value', '')
        operator = original_id_data.get('operator', '')
        if value and operator:
            if operator == 'ilike':
                wells = wells.filter(original_id__icontains=value)
            elif operator == '=':
                wells = wells.filter(original_id__iexact=value)

    # name filter
    name_data = filters.get('name', None)
    if name_data:
        value = name_data.get('value', '')
        operator = name_data.get('operator', '')
        if value and operator:
            if operator == 'ilike':
                wells = wells.filter(name__icontains=value)
            elif operator == '=':
                wells = wells.filter(name__iexact__iexact=value)

    # country filter
    country_data = filters.get('country', None)
    if country_data:
        value = country_data.get('value', '')
        operator = country_data.get('operator', '')
        if value and operator:
            if operator == 'ilike':
                wells = wells.filter(country__name__icontains=value)
            elif operator == '=':
                wells = wells.filter(country__name__iexact=value)
    return wells


@shared_task(bind=True, queue='update')
def download_well(self, user_id, download_session_id, filters=None):
    DJANGO_ROOT = os.path.dirname(
        os.path.dirname(
            os.path.dirname(os.path.abspath(__file__))
        ))

    download_session = DownloadSession.objects.get(id=download_session_id)
    user = User.objects.get(id=user_id)

    logger.debug('----- begin download  -------')
    wells = filter_wells_to_download(filters)
    total_records = wells.count()
    logger.debug('Found {} wells'.format(total_records))

    # save it to media
    unique_id = download_session.token
    folder = os.path.join(
        settings.MEDIA_ROOT, 'gwml2', 'download', str(unique_id)
    )
    if not os.path.exists(folder):
        os.makedirs(folder)

    # create file
    wells_filename = 'wells.xlsx'
    wells_file = os.path.join(folder, wells_filename)
    drilling_and_construction_filename = 'drilling_and_construction.xlsx'
    drilling_and_construction_file = os.path.join(folder, drilling_and_construction_filename)
    monitoring_filename = 'monitoring_data.xlsx'
    monitoring_file = os.path.join(folder, monitoring_filename)

    # copy template to actual folder
    copyfile(
        os.path.join(
            DJANGO_ROOT, 'gwml2', 'static', 'download_template', wells_filename),
        wells_file)
    copyfile(
        os.path.join(
            DJANGO_ROOT, 'gwml2', 'static', 'download_template', drilling_and_construction_filename),
        drilling_and_construction_file)
    copyfile(
        os.path.join(
            DJANGO_ROOT, 'gwml2', 'static', 'download_template', monitoring_filename),
        monitoring_file)

    # open sheet
    well_book = load_workbook(wells_file)
    general_information_sheet = well_book['General Information']
    hydrogeology_sheet = well_book['Hydrogeology']
    management_sheet = well_book['Management']

    drilling_and_construction_book = load_workbook(drilling_and_construction_file)
    drilling_and_construction_sheet = drilling_and_construction_book['Drilling and Construction']
    water_strike_sheet = drilling_and_construction_book['Water Strike']
    stratigraphic_sheet = drilling_and_construction_book['Stratigraphic Log']
    structure_sheet = drilling_and_construction_book['Structures']

    monitor_book = load_workbook(monitoring_file)
    level_measurement_sheet = monitor_book['Level Measurement']
    quality_measurement_sheet = monitor_book['Quality Measurement']
    yield_measurement_sheet = monitor_book['Yield Measurement']
    time.sleep(2.4)
    for index, well in enumerate(wells):
        process_percent = (index / total_records) * 100
        download_session.update_progress(process_percent)

        # check if user has view permission
        if not well.view_permission(user):
            continue

        # General Information
        general_information_sheet.append([
            well.original_id,
            well.name,
            well.status.__str__() if well.status else '',
            well.feature_type.__str__() if well.feature_type else '',
            well.purpose.__str__() if well.purpose else '',
            well.location.y,
            well.location.x,
            well.ground_surface_elevation.value if well.ground_surface_elevation else '',
            well.ground_surface_elevation.unit.name if well.ground_surface_elevation and well.ground_surface_elevation.unit else '',
            well.top_borehole_elevation.value if well.top_borehole_elevation else '',
            well.top_borehole_elevation.unit.name if well.top_borehole_elevation and well.top_borehole_elevation.unit else '',
            well.country.code if well.country else '',
            well.address,
            well.description,
        ])

        # drilling and construction
        drilling_and_construction_sheet.append([
            well.original_id,
            well.geology.total_depth.value if well.geology and well.geology.total_depth else '',
            well.geology.total_depth.unit.__str__() if well.geology and well.geology.total_depth and well.geology.total_depth.unit else '',
            well.drilling.drilling_method.__str__() if well.drilling and well.drilling.drilling_method else '',
            well.drilling.driller if well.drilling else '',
            ('Yes' if well.drilling.successful else 'No') if well.drilling and well.drilling.successful is not None else '',
            well.drilling.cause_of_failure if well.drilling else '',
            well.drilling.year_of_drilling if well.drilling else '',
            well.construction.pump_installer if well.construction else '',
            well.construction.pump_description if well.construction else '',
        ])

        if well.drilling:
            # water strike
            for water_strike in well.drilling.waterstrike_set.all():
                water_strike_sheet.append([
                    well.original_id,
                    water_strike.depth.value.value if water_strike.depth and water_strike.depth.value else '',
                    water_strike.depth.value.unit.__str__() if water_strike.depth and water_strike.depth.value and water_strike.depth.value.unit else '',
                    water_strike.depth.reference.__str__() if water_strike.depth and water_strike.depth.reference else ''
                ])

            # stratigraphic
            for stratigraphiclog in well.drilling.stratigraphiclog_set.all():
                stratigraphic_sheet.append([
                    well.original_id,
                    stratigraphiclog.reference_elevation.__str__() if stratigraphiclog.reference_elevation else '',
                    stratigraphiclog.top_depth.value if stratigraphiclog.top_depth else '',
                    stratigraphiclog.top_depth.unit.__str__() if stratigraphiclog.top_depth and stratigraphiclog.top_depth.unit else '',
                    stratigraphiclog.bottom_depth.value if stratigraphiclog.bottom_depth else '',
                    stratigraphiclog.bottom_depth.unit.__str__() if stratigraphiclog.top_depth and stratigraphiclog.bottom_depth.unit else '',
                    stratigraphiclog.material,
                    stratigraphiclog.stratigraphic_unit,
                ])
        if well.construction:
            # structures
            for structure in well.construction.constructionstructure_set.all():
                structure_sheet.append([
                    well.original_id,
                    structure.type.__str__() if structure.type else '',
                    structure.reference_elevation.__str__() if structure.reference_elevation else '',
                    structure.top_depth.value if structure.top_depth else '',
                    structure.top_depth.unit.__str__() if structure.top_depth and structure.top_depth.unit else '',
                    structure.bottom_depth.value if structure.bottom_depth else '',
                    structure.bottom_depth.unit.__str__() if structure.top_depth and structure.bottom_depth.unit else '',
                    structure.diameter.value if structure.diameter else '',
                    structure.diameter.unit.__str__() if structure.diameter and structure.diameter.unit else '',
                    structure.material,
                    structure.description
                ])

        # Hydrogeology
        hydrogeology_sheet.append([
            well.original_id,
            well.hydrogeology_parameter.aquifer_name if well.hydrogeology_parameter else '',
            well.hydrogeology_parameter.aquifer_material if well.hydrogeology_parameter else '',
            well.hydrogeology_parameter.aquifer_type.__str__() if well.hydrogeology_parameter and well.hydrogeology_parameter.aquifer_type else '',
            well.hydrogeology_parameter.aquifer_thickness.value if well.hydrogeology_parameter and well.hydrogeology_parameter.aquifer_thickness else '',
            well.hydrogeology_parameter.aquifer_thickness.unit.__str__() if well.hydrogeology_parameter and well.hydrogeology_parameter.aquifer_thickness and well.hydrogeology_parameter.aquifer_thickness.unit else '',
            well.hydrogeology_parameter.confinement.__str__() if well.hydrogeology_parameter and well.hydrogeology_parameter.confinement else '',
            well.hydrogeology_parameter.degree_of_confinement if well.hydrogeology_parameter else '',

            well.hydrogeology_parameter.pumping_test.porosity if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test else '',

            # hydraulic conductivity
            well.hydrogeology_parameter.pumping_test.hydraulic_conductivity.value if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test and well.hydrogeology_parameter.pumping_test.hydraulic_conductivity else '',
            well.hydrogeology_parameter.pumping_test.hydraulic_conductivity.unit.__str__() if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test and well.hydrogeology_parameter.pumping_test.hydraulic_conductivity and well.hydrogeology_parameter.pumping_test.hydraulic_conductivity.unit else '',

            # transmisivity
            well.hydrogeology_parameter.pumping_test.transmissivity.value if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test and well.hydrogeology_parameter.pumping_test.transmissivity else '',
            well.hydrogeology_parameter.pumping_test.transmissivity.unit.__str__() if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test and well.hydrogeology_parameter.pumping_test.transmissivity and well.hydrogeology_parameter.pumping_test.transmissivity.unit else '',

            # specific storage
            well.hydrogeology_parameter.pumping_test.specific_storage.value if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test and well.hydrogeology_parameter.pumping_test.specific_storage else '',
            well.hydrogeology_parameter.pumping_test.specific_storage.unit.__str__() if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test and well.hydrogeology_parameter.pumping_test.specific_storage and well.hydrogeology_parameter.pumping_test.specific_storage.unit else '',

            # specific capacity
            well.hydrogeology_parameter.pumping_test.specific_capacity.value if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test and well.hydrogeology_parameter.pumping_test.specific_capacity else '',
            well.hydrogeology_parameter.pumping_test.specific_capacity.unit.__str__() if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test and well.hydrogeology_parameter.pumping_test.specific_capacity and well.hydrogeology_parameter.pumping_test.specific_capacity.unit else '',

            well.hydrogeology_parameter.pumping_test.storativity if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test else '',
            well.hydrogeology_parameter.pumping_test.test_type if well.hydrogeology_parameter and well.hydrogeology_parameter.pumping_test else '',
        ])

        # Management
        management_sheet.append([
            well.original_id,
            well.organisation.name if well.organisation else '',
            well.management.manager if well.management else '',
            well.management.description if well.management else '',
            well.management.groundwater_use.__str__() if well.management and well.management.groundwater_use else '',
            well.management.number_of_users if well.management else '',

            well.management.license.number if well.management and well.management.license else '',
            well.management.license.valid_from.strftime('%Y-%m-%d') if well.management and well.management.license and well.management.license.valid_from else '',
            well.management.license.valid_until.strftime('%Y-%m-%d') if well.management and well.management.license and well.management.license.valid_until else '',
            well.management.license.description if well.management and well.management.license else '',
        ])

        # level
        for measurement in well.welllevelmeasurement_set.all():
            level_measurement_sheet.append([
                well.original_id,
                measurement.time.strftime('%Y-%m-%d %H:%M:%S %Z'),
                measurement.parameter.__str__() if measurement.parameter else '',
                measurement.value.value if measurement.value else '',
                measurement.value.unit.name if measurement.value and measurement.value.unit else '',
                measurement.methodology
            ])

        # quality
        for measurement in well.wellqualitymeasurement_set.all():
            quality_measurement_sheet.append([
                well.original_id,
                measurement.time.strftime('%Y-%m-%d %H:%M:%S %Z'),
                measurement.parameter.__str__() if measurement.parameter else '',
                measurement.value.value if measurement.value else '',
                measurement.value.unit.name if measurement.value and measurement.value.unit else '',
                measurement.methodology
            ])
        # yield
        for measurement in well.wellyieldmeasurement_set.all():
            yield_measurement_sheet.append([
                well.original_id,
                measurement.time.strftime('%Y-%m-%d %H:%M:%S %Z'),
                measurement.parameter.__str__() if measurement.parameter else '',
                measurement.value.value if measurement.value else '',
                measurement.value.unit.name if measurement.value and measurement.value.unit else '',
                measurement.methodology
            ])

    well_book.save(wells_file)
    drilling_and_construction_book.save(drilling_and_construction_file)
    monitor_book.save(monitoring_file)

    # zipping files
    zip_filename = '{}.zip'.format(str(unique_id))
    zip_file = os.path.join(
        settings.MEDIA_ROOT, 'gwml2', 'download', zip_filename
    )
    zip_file = zipfile.ZipFile(zip_file, 'w')
    zip_file.write(wells_file, wells_filename, compress_type=zipfile.ZIP_DEFLATED)
    zip_file.write(drilling_and_construction_file, drilling_and_construction_filename, compress_type=zipfile.ZIP_DEFLATED)
    zip_file.write(monitoring_file, monitoring_filename, compress_type=zipfile.ZIP_DEFLATED)
    zip_file.close()
    shutil.rmtree(folder)

    url = os.path.join(settings.MEDIA_URL, 'gwml2', 'download', zip_filename)
    download_session.file.name = os.path.join('gwml2', 'download', zip_filename)
    download_session.update_progress(100, json.dumps({'url': url}))
    return JsonResponse({'status': 'success'})
