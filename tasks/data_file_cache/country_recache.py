import json
import os
import shutil
import time
import zipfile

from django.conf import settings
from django.db.models import Q
from geonode.celery_app import app
from openpyxl import load_workbook

from gwml2.models.download_request import WELL_AND_MONITORING_DATA, GGMN
from gwml2.models.general import Country
from gwml2.models.well import Well
from gwml2.tasks.data_file_cache.base_cache import WellCacheFileBase

GWML2_FOLDER = os.getenv(
    'GWML_FOLDER', os.path.join(settings.PROJECT_ROOT, 'gwml2-file')
)
DATA_FOLDER = os.path.join(GWML2_FOLDER, 'data')
WELL_FOLDER = os.path.join(GWML2_FOLDER, 'wells-data')
DJANGO_ROOT = os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)
TEMPLATE_FOLDER = os.path.join(DJANGO_ROOT, 'static', 'download_template')


class GenerateCountryCacheFile(WellCacheFileBase):
    @property
    def country(self) -> Country:
        """Return country."""
        return self.country_data

    @property
    def folder(self) -> str:
        """Return folder.."""
        return os.path.join(DATA_FOLDER, str(self.country.code))

    def __init__(self, country):
        self.country_data = country
        self.current_time = time.time()
        self.log(f'----- Begin cache country : {country.code}  -------')

        # Prepare files
        well_folder = self.folder_by_type(WELL_AND_MONITORING_DATA)
        if not os.path.exists(well_folder):
            os.makedirs(well_folder)

        ggmn_folder = self.folder_by_type(GGMN)
        if not os.path.exists(ggmn_folder):
            os.makedirs(ggmn_folder)

        # copy files
        self.copy_template(self.wells_filename)
        self.copy_template(self.drill_filename)

        # Get data
        # Well files
        well_file = self.file_by_type(
            self.wells_filename, WELL_AND_MONITORING_DATA)
        well_book = load_workbook(well_file)
        well_ggmn_file = self.file_by_type(self.wells_filename, GGMN)
        well_ggmn_book = load_workbook(well_ggmn_file)

        # Drilling files
        drilling_file = self.file_by_type(
            self.drill_filename, WELL_AND_MONITORING_DATA)
        drilling_book = load_workbook(drilling_file)
        drilling_ggmn_file = self.file_by_type(self.drill_filename, GGMN)
        drilling_ggmn_book = load_workbook(drilling_ggmn_file)

        # Save the data
        wells = Well.objects.filter(country=self.country).order_by('id')
        for well in wells:
            self.merge_data_per_well(
                well, self.wells_filename, well_book,
                well_ggmn_book if well.number_of_measurements > 0 and well.organisation else None,
                ['General Information', 'Hydrogeology', 'Management']
            )
            self.merge_data_per_well(
                well, self.drill_filename, drilling_book,
                drilling_ggmn_book if well.number_of_measurements > 0 and well.organisation else None,
                ['Drilling and Construction', 'Water Strike',
                 'Stratigraphic Log', 'Structures']
            )

        # Save book
        well_book.save(well_file)
        well_ggmn_book.save(well_ggmn_file)
        drilling_book.save(drilling_file)
        drilling_ggmn_book.save(drilling_ggmn_file)

        # -------------------------------------------------------------------------
        # zipping files
        # -------------------------------------------------------------------------
        self.log(f'----- Finish constructing country : {country.code}  ------')
        for data_type in [WELL_AND_MONITORING_DATA, GGMN]:
            zip_filename = f'{str(self.country.code)} - {data_type}.zip'
            zip_file = os.path.join(DATA_FOLDER, zip_filename)
            if os.path.exists(zip_file):
                os.remove(zip_file)
            zip_file = zipfile.ZipFile(zip_file, 'w')

            well_file = self.file_by_type(self.wells_filename, data_type)
            zip_file.write(
                well_file, self.wells_filename,
                compress_type=zipfile.ZIP_DEFLATED)

            drill_file = self.file_by_type(self.drill_filename, data_type)
            zip_file.write(
                drill_file,
                self.drill_filename,
                compress_type=zipfile.ZIP_DEFLATED)

            for well in wells:
                well_folder = os.path.join(WELL_FOLDER, f'{well.id}')
                measurement_file = os.path.join(
                    well_folder, self.monitor_filename
                )
                if os.path.exists(measurement_file):
                    zip_file.write(
                        measurement_file,
                        f'monitoring/{well.original_id} ({well.id}).xlsx',
                        compress_type=zipfile.ZIP_DEFLATED
                    )

            zip_file.close()
        shutil.rmtree(self.folder)
        self.log(f'----- Finish zipping : {country.code}  -------')

    def merge_data_per_well(
            self, well, filename, well_book, ggmn_book, sheets
    ):
        """Merge data per well.."""
        well_folder = os.path.join(WELL_FOLDER, f'{well.id}')
        for sheetname in sheets:
            self.merge_data_between_sheets(
                os.path.join(well_folder, filename),
                well_book, ggmn_book, sheetname
            )

    def merge_data_between_sheets(
            self, source_file, target_book, target_book_2, sheetname
    ):
        """Merge data between sheets"""
        if not os.path.exists(source_file) or not target_book:
            return
        source_file = os.path.join(source_file, sheetname + '.json')
        data = []
        if os.path.exists(source_file):
            _file = open(source_file, "r")
            data = json.loads(_file.read())

        # Target book 1
        target_sheet = target_book[sheetname]

        # Target book 2
        target_sheet_2 = None
        if target_book_2:
            target_sheet_2 = target_book_2[sheetname]

        # Append data from source
        for row in data:
            target_sheet.append(row)
            if target_book_2:
                target_sheet_2.append(row)


@app.task(
    bind=True,
    name='gwml2.tasks.well.generate_data_country_cache'
)
def generate_data_country_cache(self, country_code: str):
    try:
        country = Country.objects.get(Q(code__iexact=country_code))
        GenerateCountryCacheFile(country)
    except Country.DoesNotExist:
        print('Country not found')
