import json
import os
import shutil
import time
import zipfile

from celery.utils.log import get_task_logger
from django.conf import settings
from openpyxl import load_workbook

from gwml2.models.download_request import WELL_AND_MONITORING_DATA, GGMN
from gwml2.models.term import TermFeatureType
from gwml2.models.well import Organisation
from gwml2.terms import SheetName
from gwml2.utilities import xlsx_to_ods

GWML2_FOLDER = settings.GWML2_FOLDER
WELL_FOLDER = os.path.join(GWML2_FOLDER, 'wells-data')
DJANGO_ROOT = os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)
TEMPLATE_FOLDER = os.path.join(DJANGO_ROOT, 'static', 'download_template')

logger = get_task_logger(__name__)


def zipdir(path, ziph):
    """Zip directory"""
    for root, dirs, files in os.walk(path):
        for file in files:
            ziph.write(
                os.path.join(root, file),
                os.path.relpath(
                    os.path.join(root, file), os.path.join(path, '..')
                )
            )


def get_data(id, cache, Term):
    """Get data that on cache or not."""
    if not id:
        return ''
    if id not in cache:
        try:
            value = Term.objects.get(id=id).__str__()
        except TermFeatureType.DoesNotExist:
            value = ''
        cache[id] = value
        return value
    return cache[id]


class WellCacheFileBase(object):
    current_time = None
    wells_filename = 'wells.xlsx'
    drill_filename = 'drilling_and_construction.xlsx'
    monitor_filename = 'monitoring_data.ods'

    # cache
    feature_types = {}
    purposes = {}
    status = {}
    units = {}
    drillings = {}
    reference_elevations = {}
    structures_types = {}
    aquifer_types = {}
    confinements = {}
    organisations = {}
    groundwater_uses = {}
    measurement_parameters = {}

    @property
    def folder(self) -> str:
        """Return temporary working directory."""
        raise NotImplementedError

    def folder_by_type(self, data_type):
        """Return folder by type."""
        return os.path.join(self.folder, data_type)

    def file_by_type(self, filename, data_type):
        """Get file of on folder."""
        return os.path.join(self.folder_by_type(data_type), filename)

    def copy_template(self, filename):
        """Copy template."""
        well_file = self.file_by_type(filename, WELL_AND_MONITORING_DATA)
        ggmn_file = self.file_by_type(filename, GGMN)

        # Delete existed file
        if os.path.exists(well_file):
            os.remove(well_file)
        if os.path.exists(ggmn_file):
            os.remove(ggmn_file)

        # Copy the template
        shutil.copyfile(os.path.join(TEMPLATE_FOLDER, filename), well_file)
        shutil.copyfile(os.path.join(TEMPLATE_FOLDER, filename), ggmn_file)

    def log(self, text):
        """Print time."""
        new_time = time.time()
        print(f'{text} - {(new_time - self.current_time)} seconds')
        logger.debug(f'{text} - {(new_time - self.current_time)} seconds')
        self.current_time = new_time


class WellCacheZipFileBase(WellCacheFileBase):
    data_folder = None

    @property
    def cache_type(self) -> str:
        """Return type of cache (country or organisation)."""
        raise NotImplementedError

    @property
    def cache_name(self) -> str:
        """Return name of cache object."""
        raise NotImplementedError

    @property
    def folder(self) -> str:
        """Return folder.."""
        return os.path.join(self.data_folder, self.cache_name)

    def zip_file_path(self, data_type) -> str:
        """Return path to the final zip file."""
        zip_filename = f'{self.cache_name} - {data_type}.zip'
        return os.path.join(self.data_folder, zip_filename)

    def get_well_queryset(self):
        """Return queryset for wells."""
        raise NotImplementedError

    def merge_data_per_well(
            self, well, filename, well_book, ggmn_book, sheets
    ):
        """Merge data per well.."""
        well_folder = os.path.join(WELL_FOLDER, f'{well.id}')
        for sheetname in sheets:
            self.merge_data_between_sheets(
                os.path.join(well_folder, filename),
                well_book, ggmn_book, sheetname
            )
        if ggmn_book:
            ggmn_book.active = 0
        if well_book:
            well_book.active = 0

    def merge_data_between_sheets(
            self, source_file, target_book, target_book_2, sheetname
    ):
        """Merge data between sheets"""
        if (
                not os.path.exists(source_file)
                or (not target_book and not target_book_2)
        ):
            return
        source_file = os.path.join(source_file, sheetname + '.json')
        data = []
        if os.path.exists(source_file):
            _file = open(source_file, "r")
            data = json.loads(_file.read())

        # Target book 1
        target_sheet = None
        if target_book:
            target_sheet = target_book[sheetname]

        # Target book 2
        target_sheet_2 = None
        if target_book_2:
            target_sheet_2 = target_book_2[sheetname]

        # Append data from source
        for row in data:
            if target_book:
                target_sheet.append(row)
            if target_book_2:
                target_sheet_2.append(row)

    def clear_folder(self):
        if os.path.exists(self.folder):
            try:
                shutil.rmtree(self.folder)
            except FileNotFoundError:
                pass

    def zip_excel_to_ods(self, zip_file, filename, data_type):
        """Zip excel to ods."""
        well_file = self.file_by_type(
            filename, data_type
        )
        xlsx_to_ods(well_file)
        zip_file.write(
            well_file.replace('.xlsx', '.ods'),
            filename.replace('.xlsx', '.ods'),
            compress_type=zipfile.ZIP_DEFLATED
        )

    def run(self):
        self.current_time = time.time()
        self.log(
            f'----- Begin cache {self.cache_type}: {self.cache_name} '
            '-------'
        )
        # clear everything before starts
        self.clear_folder()

        # Prepare files
        well_folder = self.folder_by_type(WELL_AND_MONITORING_DATA)
        if not os.path.exists(well_folder):
            os.makedirs(well_folder)

        ggmn_folder = self.folder_by_type(GGMN)
        if not os.path.exists(ggmn_folder):
            os.makedirs(ggmn_folder)

        # copy files
        self.copy_template(self.wells_filename)
        self.copy_template(self.drill_filename)

        # Get data
        # Well files
        well_file = self.file_by_type(
            self.wells_filename, WELL_AND_MONITORING_DATA
        )
        well_book = load_workbook(well_file)
        well_ggmn_file = self.file_by_type(self.wells_filename, GGMN)
        well_ggmn_book = load_workbook(well_ggmn_file)

        # Drilling files
        drilling_file = self.file_by_type(
            self.drill_filename, WELL_AND_MONITORING_DATA)
        drilling_book = load_workbook(drilling_file)
        drilling_ggmn_file = self.file_by_type(self.drill_filename, GGMN)
        drilling_ggmn_book = load_workbook(drilling_ggmn_file)

        # Get list of ggmn organisation
        ggmn_organisations_list = Organisation.ggmn_organisations()

        # Save the data
        wells = self.get_well_queryset()
        for well in wells:
            is_ggmn = well.is_ggmn(
                ggmn_organisations_list
            ) and well.organisation

            self.merge_data_per_well(
                well, self.wells_filename,
                well_book if not is_ggmn else None,
                well_ggmn_book if is_ggmn else None,
                ['General Information', 'Hydrogeology', 'Management']
            )
            self.merge_data_per_well(
                well, self.drill_filename,
                drilling_book if not is_ggmn else None,
                drilling_ggmn_book if is_ggmn else None,
                [
                    SheetName.drilling_and_construction,
                    'Water Strike', 'Stratigraphic Log',
                    SheetName.structure
                ]
            )

        # Save book
        well_book.save(well_file)
        well_ggmn_book.save(well_ggmn_file)
        drilling_book.save(drilling_file)
        drilling_ggmn_book.save(drilling_ggmn_file)

        # -------------------------------------------------------------------------
        # zipping files
        # -------------------------------------------------------------------------
        self.log(
            f'----- Finish constructing {self.cache_type}: {self.cache_name} '
            '------'
        )
        for data_type in [WELL_AND_MONITORING_DATA, GGMN]:
            # Get the file path
            zip_filepath = self.zip_file_path(data_type)
            if os.path.exists(zip_filepath):
                os.remove(zip_filepath)
            zip_file = None

            original_ids_found = {}
            wells = self.get_well_queryset()
            for well in wells:
                is_ggmn = well.is_ggmn(
                    ggmn_organisations_list
                ) and well.organisation

                if data_type == GGMN and not is_ggmn:
                    continue
                if data_type == WELL_AND_MONITORING_DATA and is_ggmn:
                    continue

                if not zip_file:
                    zip_file = zipfile.ZipFile(zip_filepath, 'w')

                    self.zip_excel_to_ods(
                        zip_file, self.wells_filename, data_type
                    )
                    self.zip_excel_to_ods(
                        zip_file, self.drill_filename, data_type
                    )

                if well.number_of_measurements == 0:
                    continue

                well_folder = os.path.join(WELL_FOLDER, f'{well.id}')
                measurement_file = os.path.join(
                    well_folder, self.monitor_filename
                )
                if os.path.exists(measurement_file):
                    original_id = well.original_id
                    try:
                        _filename = (
                            f'monitoring/{original_id} '
                            f'({original_ids_found[original_id] + 1}).ods'
                        )
                        continue
                    except KeyError:
                        _filename = f'monitoring/{original_id}.ods'
                        original_ids_found[original_id] = 0

                    zip_file.write(
                        measurement_file,
                        _filename,
                        compress_type=zipfile.ZIP_DEFLATED
                    )
            if zip_file:
                zip_file.close()
        self.log(
            f'----- Finish zipping {self.cache_type}: {self.cache_name} '
            '-------'
        )

        # clear temp directory
        self.clear_folder()
