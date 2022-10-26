import os
import shutil
import time
import zipfile
from shutil import copyfile

from celery.utils.log import get_task_logger
from django.conf import settings
from django.db.models import Q
from geonode.celery_app import app
from openpyxl import load_workbook

logger = get_task_logger(__name__)

from gwml2.models.general import Country, Unit
from gwml2.models.term import (
    TermFeatureType, TermWellPurpose, TermWellStatus, TermDrillingMethod,
    TermReferenceElevationType, TermConstructionStructureType,
    TermAquiferType, TermConfinement, TermGroundwaterUse
)
from gwml2.models.well import Well
from gwml2.models.well_management.organisation import Organisation
from gwml2.models.term_measurement_parameter import TermMeasurementParameter
from gwml2.models.download_request import WELL_AND_MONITORING_DATA, GGMN

GWML2_FOLDER = os.getenv(
    'GWML_FOLDER', os.path.join(settings.PROJECT_ROOT, 'gwml2-file')
)
DATA_FOLDER = os.path.join(GWML2_FOLDER, 'data')
DJANGO_ROOT = os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)
TEMPLATE_FOLDER = os.path.join(
    DJANGO_ROOT, 'gwml2', 'static', 'download_template'
)


class GenerateDownloadFile(object):
    current_time = None
    wells_filename = 'wells.xlsx'
    drill_filename = 'drilling_and_construction.xlsx'
    monitor_filename = 'monitoring_data.xlsx'
    error_filename = 'error'

    # cache
    feature_types = {}
    purposes = {}
    status = {}
    units = {}
    drillings = {}
    reference_elevations = {}
    structures_types = {}
    aquifer_types = {}
    confinements = {}
    organisations = {}
    groundwater_uses = {}
    measurement_parameters = {}

    def folder_by_type(self, data_type):
        """Return folder by type."""
        return os.path.join(self.folder, data_type)

    def file_by_type(self, filename, data_type):
        """Get file of on folder."""
        return os.path.join(self.folder_by_type(data_type), filename)

    def copy_template(self, filename):
        """Copy template."""
        copyfile(
            os.path.join(TEMPLATE_FOLDER, filename),
            self.file_by_type(filename, WELL_AND_MONITORING_DATA)
        )
        copyfile(
            os.path.join(TEMPLATE_FOLDER, filename),
            self.file_by_type(filename, GGMN)
        )

    def __init__(self, country):
        print(f'----- begin download : {country.code}  -------')
        self.current_time = time.time()

        # Prepare files
        self.country = country
        self.folder = os.path.join(DATA_FOLDER, str(country.code))
        if os.path.exists(self.folder):
            shutil.rmtree(self.folder)
        os.makedirs(os.path.join(self.folder, WELL_AND_MONITORING_DATA))
        os.makedirs(os.path.join(self.folder, GGMN))

        self.error_file = os.path.join(self.folder, self.error_filename)

        # copy files
        self.copy_template(self.wells_filename)
        self.copy_template(self.drill_filename)
        self.copy_template(self.monitor_filename)
        self.run_wells()

    def log(self, text):
        """ Print time """
        new_time = time.time()
        print(f'{(new_time - self.current_time)} seconds : {text}')
        self.current_time = new_time

    def get_data(self, id, cache, Term):
        """ Get data that on cache or not. """
        if not id:
            return ''
        if id not in cache:
            try:
                value = Term.objects.get(id=id).__str__()
            except TermFeatureType.DoesNotExist:
                value = ''
            cache[id] = value
            return value
        return cache[id]

    def return_well_and_ggmn_files(self, well, filename):
        """Return well and ggmn files"""
        well_file = self.file_by_type(filename, WELL_AND_MONITORING_DATA)
        ggmn_file = None
        if well.number_of_measurements > 0 and well.organisation:
            ggmn_file = self.file_by_type(filename, GGMN)
        return well_file, ggmn_file

    def run_wells(self):
        """ Run wells """
        wells = Well.objects.all()
        if self.country:
            wells = wells.filter(country=self.country)

        total_records = wells.count()
        self.log(f'Found {total_records} wells')

        # Start check wells
        self.log('Start')
        is_error = False
        for index, well in enumerate(wells):
            try:
                print(f'Progress {index}/{total_records} : {well.original_id}')
                wells_file, wells_ggmn_file = self.return_well_and_ggmn_files(
                    well, self.wells_filename
                )
                well_book = load_workbook(wells_file)
                well_book_ggmn = load_workbook(
                    wells_ggmn_file
                ) if wells_ggmn_file else None

                # General information
                self.general_information(well_book, well_book_ggmn, well)
                self.hydrogeology(well_book, well_book_ggmn, well)
                self.management(well_book, well_book_ggmn, well)
                well_book.save(wells_file)
                if well_book_ggmn:
                    well_book_ggmn.save(wells_ggmn_file)

                # Drill
                drill_file, drill_ggmn_file = self.return_well_and_ggmn_files(
                    well, self.drill_filename
                )
                drilling_book = load_workbook(drill_file)
                drilling_ggmn_book = load_workbook(
                    drill_ggmn_file
                ) if drill_ggmn_file else None
                self.drilling_and_construction(
                    drilling_book, drilling_ggmn_book, well
                )
                drilling_book.save(drill_file)
                if drilling_ggmn_book:
                    drilling_ggmn_book.save(drill_ggmn_file)

                # Monitor
                monitor_file, monitor_ggmn_file = self.return_well_and_ggmn_files(
                    well, self.monitor_filename
                )
                monitor_book = load_workbook(monitor_file)
                monitor_ggmn_book = load_workbook(
                    monitor_ggmn_file
                ) if monitor_ggmn_file else None
                self.measurements(monitor_book, monitor_ggmn_book, well)
                monitor_book.save(monitor_file)
                if monitor_ggmn_book:
                    monitor_ggmn_book.save(monitor_ggmn_file)
            except Exception as e:
                error_file = open(self.error_file, 'a')
                error = f'{well.original_id} : {e}'
                print(error)
                error_file.write(error + '\n')
                error_file.close()
                is_error = True

        self.log('Finish')
        if is_error:
            self.log('Finish with error')
            return

        # -------------------------------------------------------------------------
        # zipping files
        # -------------------------------------------------------------------------
        for data_type in [WELL_AND_MONITORING_DATA, GGMN]:
            zip_filename = f'{str(self.country.code)} - {data_type}.zip'
            zip_file = os.path.join(DATA_FOLDER, zip_filename)
            if os.path.exists(zip_file):
                os.remove(zip_file)

            well_file = self.file_by_type(self.wells_filename, data_type)
            drill_file = self.file_by_type(self.drill_filename, data_type)
            monitor_file = self.file_by_type(self.monitor_filename, data_type)
            zip_file = zipfile.ZipFile(zip_file, 'w')
            zip_file.write(
                well_file, self.wells_filename,
                compress_type=zipfile.ZIP_DEFLATED)
            zip_file.write(
                drill_file,
                self.drill_filename,
                compress_type=zipfile.ZIP_DEFLATED)

            zip_file.write(
                monitor_file, self.monitor_filename,
                compress_type=zipfile.ZIP_DEFLATED)

            zip_file.close()
        shutil.rmtree(self.folder)

    def general_information(self, book, ggmn_book, well):
        """ General Information of well"""
        sheetname = 'General Information'
        data = [
            well.original_id,
            well.name,
            self.get_data(
                well.feature_type_id, self.feature_types, TermFeatureType
            ),
            self.get_data(well.purpose_id, self.purposes, TermWellPurpose),
            self.get_data(well.status_id, self.status, TermWellStatus),
            well.description,
            well.location.y,
            well.location.x,

            # Ground surface elevation
            well.ground_surface_elevation.value if well.ground_surface_elevation else '',
            self.get_data(
                well.ground_surface_elevation.unit_id, self.units, Unit)
            if well.ground_surface_elevation else '',

            # Top borehole elevation
            well.top_borehole_elevation.value if well.top_borehole_elevation else '',
            self.get_data(
                well.top_borehole_elevation.unit_id, self.units, Unit)
            if well.top_borehole_elevation else '',
            self.country.code,
            well.address,
        ]
        sheet = book[sheetname]
        sheet.append(data)

        # If has ggmn book
        if ggmn_book:
            sheet = ggmn_book[sheetname]
            sheet.append(data)

    def drilling_and_construction(self, book, ggmn_book, well):
        """ Drilling and construction of well"""
        geology = well.geology if well.geology else None
        drilling = well.drilling if well.drilling else None
        construction = well.construction if well.construction else None

        data = [
            well.original_id,

            # Total depth
            geology.total_depth.value if geology and geology.total_depth else '',
            self.get_data(geology.total_depth.unit_id, self.units, Unit)
            if geology and geology.total_depth else '',

            # Drilling
            self.get_data(
                drilling.drilling_method_id, self.drillings, TermDrillingMethod
            )
            if drilling and drilling.drilling_method else '',
            drilling.driller if drilling else '',
            (
                'Yes' if drilling.successful else 'No'
            ) if drilling and drilling.successful is not None else '',
            drilling.cause_of_failure if drilling else '',
            drilling.year_of_drilling if drilling else '',
            construction.pump_installer if construction else '',
            construction.pump_description if construction else '',
        ]
        sheetname = 'Drilling and Construction'
        sheet = book[sheetname]
        sheet.append(data)

        # If has ggmn book
        if ggmn_book:
            sheet = ggmn_book[sheetname]
            sheet.append(data)

        # --------------------------------------------------------------------------
        # For drilling data
        if drilling:
            # water strike
            sheetname = 'Water Strike'
            for water_strike in well.drilling.waterstrike_set.all():
                depth = water_strike.depth
                data = [
                    well.original_id,

                    # Depth
                    depth.value if depth else '',
                    self.get_data(
                        depth.unit_id, self.units, Unit
                    ) if depth else '',
                    self.get_data(
                        water_strike.reference_elevation_id,
                        self.reference_elevations, TermReferenceElevationType
                    ),
                ]
                sheet = book[sheetname]
                sheet.append(data)

                # If has ggmn book
                if ggmn_book:
                    sheet = ggmn_book[sheetname]
                    sheet.append(data)

            # stratigraphic
            sheetname = 'Stratigraphic Log'
            for log in well.drilling.stratigraphiclog_set.all():
                top_depth = log.top_depth
                bottom_depth = log.bottom_depth
                data = [
                    well.original_id,

                    # Reference elevation
                    self.get_data(
                        log.reference_elevation_id, self.reference_elevations,
                        TermReferenceElevationType
                    ),

                    # Top depth
                    top_depth.value if top_depth else '',
                    self.get_data(top_depth.unit_id, self.units, Unit)
                    if top_depth else '',

                    # Bottom depth
                    bottom_depth.value if bottom_depth else '',
                    self.get_data(bottom_depth.unit_id, self.units, Unit)
                    if bottom_depth else '',
                    log.material,
                    log.stratigraphic_unit,
                ]
                sheet = book[sheetname]
                sheet.append(data)

                # If has ggmn book
                if ggmn_book:
                    sheet = ggmn_book[sheetname]
                    sheet.append(data)
        # --------------------------------------------------------------------------
        # For Construction Data
        if construction:
            sheetname = 'Structures'
            for structure in well.construction.constructionstructure_set.all():
                top_depth = structure.top_depth
                bottom_depth = structure.bottom_depth
                diameter = structure.diameter
                data = [
                    well.original_id,

                    # Structure
                    self.get_data(
                        structure.type_id, self.structures_types,
                        TermConstructionStructureType
                    ),
                    self.get_data(
                        structure.reference_elevation_id,
                        self.reference_elevations,
                        TermReferenceElevationType
                    ),

                    # Top depth
                    top_depth.value if top_depth else '',
                    self.get_data(top_depth.unit_id, self.units, Unit)
                    if top_depth else '',

                    # Bottom depth
                    bottom_depth.value if bottom_depth else '',
                    self.get_data(bottom_depth.unit_id, self.units, Unit)
                    if bottom_depth else '',

                    # Diameter
                    diameter.value if diameter else '',
                    self.get_data(diameter.unit_id, self.units, Unit)
                    if diameter else '',
                    structure.material,
                    structure.description
                ]
                sheet = book[sheetname]
                sheet.append(data)

                # If has ggmn book
                if ggmn_book:
                    sheet = ggmn_book[sheetname]
                    sheet.append(data)

    def hydrogeology(self, book, ggmn_book, well):
        """ Hydrogeology of well"""
        hydrogeology = well.hydrogeology_parameter
        pumping_test = hydrogeology.pumping_test if hydrogeology else None
        hydraulic_conductivity = pumping_test.hydraulic_conductivity if pumping_test else None
        transmissivity = pumping_test.transmissivity if pumping_test else None
        specific_storage = pumping_test.specific_storage if pumping_test else None
        specific_capacity = pumping_test.specific_capacity if pumping_test else None
        storativity = pumping_test.storativity if pumping_test else None

        data = [
            well.original_id,
            hydrogeology.aquifer_name if hydrogeology else '',
            hydrogeology.aquifer_material if hydrogeology else '',

            # Aquifer type
            self.get_data(
                hydrogeology.aquifer_type_id, self.aquifer_types,
                TermAquiferType
            )
            if hydrogeology else '',

            hydrogeology.aquifer_thickness
            if hydrogeology and hydrogeology.aquifer_thickness else '',

            # Aquifer confinement
            self.get_data(
                hydrogeology.confinement_id, self.confinements,
                TermConfinement
            )
            if hydrogeology else '',

            hydrogeology.degree_of_confinement if hydrogeology else '',

            # Pumping test
            pumping_test.porosity if pumping_test else '',

            hydraulic_conductivity.value if hydraulic_conductivity else '',
            self.get_data(hydraulic_conductivity.unit_id, self.units, Unit)
            if hydraulic_conductivity else '',

            # transmisivity
            transmissivity.value if transmissivity else '',
            self.get_data(transmissivity.unit_id, self.units, Unit)
            if transmissivity else '',

            # specific storage
            specific_storage.value if specific_storage else '',
            self.get_data(specific_storage.unit_id, self.units, Unit)
            if specific_storage else '',

            # specific capacity
            specific_capacity.value if specific_capacity else '',
            self.get_data(specific_capacity.unit_id, self.units, Unit)
            if specific_capacity else '',

            # specific capacity
            storativity.value if storativity else '',
            self.get_data(storativity.unit_id, self.units, Unit)
            if storativity else '',

            pumping_test.test_type if pumping_test else '',
        ]
        sheetname = 'Hydrogeology'
        sheet = book[sheetname]
        sheet.append(data)

        # If has ggmn book
        if ggmn_book:
            sheet = ggmn_book[sheetname]
            sheet.append(data)

    def management(self, book, ggmn_book, well):
        """ Management of well"""
        management = well.management
        license = management.license if management else None
        data = [
            well.original_id,
            self.get_data(
                well.organisation_id, self.organisations, Organisation
            ),

            # management
            management.manager if management else '',
            management.description if management else '',
            self.get_data(
                management.groundwater_use_id, self.groundwater_uses,
                TermGroundwaterUse
            ) if management else '',
            management.number_of_users if management else '',

            license.number if license else '',
            license.valid_from.strftime('%Y-%m-%d')
            if license and license.valid_from else '',
            license.valid_until.strftime('%Y-%m-%d')
            if management and license and license.valid_until else '',
            license.description if license else '',
        ]

        sheetname = 'Management'
        sheet = book[sheetname]
        sheet.append(data)

        # If has ggmn book
        if ggmn_book:
            sheet = ggmn_book[sheetname]
            sheet.append(data)

    def measurement_data(self, sheets, ggmn_sheet, measurements, original_id):
        """ Measurements of well """
        for measurement in measurements:
            value = measurement.value
            data = [
                original_id,

                # measurement
                measurement.time.strftime('%Y-%m-%d %H:%M:%S %Z'),
                self.get_data(
                    measurement.parameter_id, self.measurement_parameters,
                    TermMeasurementParameter
                ),

                # value
                value.value if value else '',
                self.get_data(value.unit_id, self.units, Unit)
                if value else '',
                measurement.methodology
            ]
            sheets.append(data)
            if ggmn_sheet:
                ggmn_sheet.append(data)

    def measurements(self, book, ggmn_book, well):
        """ Measurements of well """
        self.measurement_data(
            book['Groundwater Level'],
            ggmn_book['Groundwater Level'] if ggmn_book else None,
            well.welllevelmeasurement_set.all(),
            well.original_id
        )
        self.measurement_data(
            book['Groundwater Quality'],
            ggmn_book['Groundwater Quality'] if ggmn_book else None,
            well.wellqualitymeasurement_set.all(),
            well.original_id
        )
        self.measurement_data(
            book['Abstraction-Discharge'],
            ggmn_book['Abstraction-Discharge'] if ggmn_book else None,
            well.wellyieldmeasurement_set.all(),
            well.original_id
        )


@app.task(
    bind=True,
    name='gwml2.tasks.well.generate_downloadable_file_cache'
)
def generate_downloadable_file_cache(
        self, country: str = None, is_from: bool = False):
    countries = Country.objects.order_by('name')
    if country:
        try:
            country = countries.get(Q(code__iexact=country))
            if not is_from:
                GenerateDownloadFile(country)
            else:
                for country in countries.filter(name__gte=country.name):
                    GenerateDownloadFile(country)

        except Country.DoesNotExist:
            print('Country not found')
    else:
        for country in countries:
            GenerateDownloadFile(country)
