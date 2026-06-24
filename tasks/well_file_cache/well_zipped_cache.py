import json
import os
import shutil
import time
import zipfile

from celery.utils.log import get_task_logger
from openpyxl import load_workbook

from gwml2.terms import SheetName
from gwml2.utilities import xlsx_to_ods

DJANGO_ROOT = os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)
TEMPLATE_FOLDER = os.path.join(DJANGO_ROOT, 'static', 'download_template')

logger = get_task_logger(__name__)


class WellZippedCache(object):
    current_time = None
    wells_filename = 'wells.xlsx'
    drill_filename = 'drilling_and_construction.xlsx'
    monitor_filename = 'monitoring_data.ods'

    @property
    def data_folder(self) -> str:
        """Return data folder path."""
        raise NotImplementedError

    def filepath(self, filename):
        """Get file on folder."""
        return os.path.join(self.folder, filename)

    def clean(self):
        """Remove all temporary files and folders."""
        if os.path.exists(self.folder):
            try:
                shutil.rmtree(self.folder)
            except FileNotFoundError:
                pass

    def copy_template(self, filename):
        """Copy template."""
        well_file = self.filepath(filename)
        shutil.copyfile(os.path.join(TEMPLATE_FOLDER, filename), well_file)

    def log(self, text):
        """Print time."""
        new_time = time.time()
        print(f'{text} - {(new_time - self.current_time)} seconds')
        logger.debug(f'{text} - {(new_time - self.current_time)} seconds')
        self.current_time = new_time

    @property
    def cache_name(self) -> str:
        """Return name of cache object."""
        raise NotImplementedError

    @property
    def well_queryset(self):
        """Return queryset for wells."""
        raise NotImplementedError

    @property
    def folder(self) -> str:
        """Return folder."""
        return os.path.join(self.data_folder, self.cache_name)

    @property
    def zip_file_path(self) -> str:
        """Return path to the final zip file."""
        return os.path.join(self.data_folder, f'{self.cache_name}.zip')

    def merge_data_per_well(
            self, well, filename, well_book, sheets
    ):
        """Merge data per well.."""
        well_folder = well.data_cache_folder

        # Load data of well
        well_data = None
        data_file = os.path.join(well_folder, 'data.json')
        if os.path.exists(data_file):
            try:
                with open(data_file, 'r') as f:
                    well_data = json.load(f)
            except (json.JSONDecodeError, OSError):
                pass
        for sheetname in sheets:
            self.merge_data_between_sheets(
                well_data,
                os.path.join(well_folder, filename), well_book, sheetname
            )
        if well_book:
            well_book.active = 0

    def merge_data_between_sheets(
            self, well_data, source_folder, target_book, sheetname
    ):
        """Merge data between sheets"""
        target_column_number = SheetName().get_column_size(
            sheet_name=sheetname
        )

        # -------------------------------
        # This is new approach
        # -------------------------------
        if well_data:
            try:
                data = well_data[sheetname]
            except KeyError:
                return
        else:
            # -------------------------------
            # This is old approach
            # -------------------------------
            if not os.path.exists(source_folder) or not target_book:
                return
            source_file = os.path.join(source_folder, f'{sheetname}.json')
            if not os.path.exists(source_file):
                return
            with open(source_file, 'r') as f:
                try:
                    data = json.load(f)
                except json.JSONDecodeError:
                    return

        # Target book
        target_sheet = None
        if target_book:
            target_sheet = target_book[sheetname]

        # Append data from source
        for row in data:
            if len(row) != target_column_number:
                continue

            if target_book:
                target_sheet.append(row)

    def zip_excel_to_ods(self, zip_file, filename):
        """Zip excel to ods."""
        well_file = self.filepath(filename)
        xlsx_to_ods(well_file)
        zip_file.write(
            well_file.replace('.xlsx', '.ods'),
            filename.replace('.xlsx', '.ods'),
            compress_type=zipfile.ZIP_DEFLATED
        )

    def run(self, post_function=None):
        self.current_time = time.time()
        self.log(
            f'----- Begin cache {self.cache_name} -------'
        )
        # clear everything before starts
        self.clean()

        # Prepare files
        well_folder = self.folder
        if not os.path.exists(well_folder):
            os.makedirs(well_folder)

        # copy files
        self.copy_template(self.wells_filename)
        self.copy_template(self.drill_filename)

        # Get data
        # Well files
        well_file = self.filepath(self.wells_filename)
        well_book = load_workbook(well_file)
        drilling_file = self.filepath(self.drill_filename)
        drilling_book = load_workbook(drilling_file)

        # Save the data
        wells = self.well_queryset
        for well in wells:
            self.merge_data_per_well(
                well, self.wells_filename, well_book,
                [
                    SheetName.general_information,
                    SheetName.hydrogeology,
                    SheetName.management
                ]
            )
            self.merge_data_per_well(
                well, self.drill_filename, drilling_book,
                [
                    SheetName.drilling_and_construction,
                    SheetName.water_strike,
                    SheetName.stratigraphic_log,
                    SheetName.structure
                ]
            )

        # Save book
        well_book.save(well_file)
        well_book.close()
        drilling_book.save(drilling_file)
        drilling_book.close()
        # -------------------------------------------------------------------------
        # zipping files
        # -------------------------------
        self.log(
            f'-- Finish constructing {self.cache_name} ----'
        )
        # Get the file path
        zip_filepath = self.zip_file_path
        if os.path.exists(zip_filepath):
            os.remove(zip_filepath)

        zip_file = None
        try:
            original_ids_found = {}
            wells = self.well_queryset
            for well in wells:
                if not zip_file:
                    zip_file = zipfile.ZipFile(zip_filepath, 'w')

                    self.zip_excel_to_ods(
                        zip_file, self.wells_filename
                    )
                    self.zip_excel_to_ods(
                        zip_file, self.drill_filename
                    )

                # Save measurements
                if well.number_of_measurements == 0:
                    continue

                well_folder = well.data_cache_folder
                measurement_file = os.path.join(
                    well_folder, self.monitor_filename
                )
                if os.path.exists(measurement_file):
                    original_id = well.original_id
                    try:
                        _filename = (
                            f'monitoring/{original_id} '
                            f'({original_ids_found[original_id] + 1}).ods'
                        )
                    except KeyError:
                        _filename = f'monitoring/{original_id}.ods'
                        original_ids_found[original_id] = 0

                    zip_file.write(
                        measurement_file,
                        _filename,
                        compress_type=zipfile.ZIP_DEFLATED
                    )
            if post_function:
                post_function(zip_file)
        except Exception as e:
            raise e
        finally:
            if zip_file:
                zip_file.close()
        self.log(
            f'---- Finish zipping {self.cache_name} ------'
        )

        # clear temp directory
        self.clean()


class WellZippedCacheByIds(WellZippedCache):
    """Cache wells by ids."""

    data_folder = None
    cache_name = None
    well_queryset = None

    def __init__(self, data_folder, cache_name, well_queryset):
        self.data_folder = data_folder
        self.cache_name = cache_name
        self.well_queryset = well_queryset
